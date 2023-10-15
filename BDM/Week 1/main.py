import openpyxl
db = "<YOUR EXCEL FILE HERE>.xlsx"
doc = openpyxl.load_workbook(db)

final = ''

grades = {'A':	'Distinction',
'B':	'First Class',
'C':	'Second Class',
'D':	'Third Class',
'E':	'Pass',
'F':	'Fail'}

def grade(mark):
    if mark >=90:
        return 'A'
    elif mark >= 75:
        return 'B'
    elif mark >= 60:
        return 'C'
    elif mark >= 50:
        return 'D'
    elif mark >= 35:
        return 'E'
    return 'F'
def avg(l):
    return round(sum(l)/len(l),2)

sheet = doc.active

mrow, mcol = sheet.max_row+1,sheet.max_column

depts =  [sheet.cell(row=i,column=2).value for i in range(2,mrow)]

math,phy,chem = [sheet.cell(row=i,column=3).value for i in range(2,mrow)],[sheet.cell(row=i,column=4).value for i in range(2,mrow)],[sheet.cell(row=i,column=5).value for i in range(2,mrow)]

for i in range(2,mrow): sheet.cell(row=i,column=7).value = 4 * sheet.cell(row=i,column=6).value

langadj = [sheet.cell(row=i,column=7).value for i in range(2,mrow)]

tavg = [(i+j+k+l)/4 for i,j,k,l in zip(math,phy,chem,langadj)]

final += f"Q1,{max(math)}\n"
final += f'Q2,{min(phy)}\n'
final += f'Q3,{avg(chem)}\n'
final += f'Q4,{avg(langadj)}\n'
final += f'Q5,{len([i for i,j in zip(depts,tavg) if i=="MECH" and grade(j)=="D"])}\n'
final += f'Q6,{len([i for i,j in zip(depts,tavg) if i=="CS" and grade(j)=="C"])}'

with open("output.csv","w") as f:
    f.write(final)





